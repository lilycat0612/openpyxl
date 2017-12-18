# -*- coding:utf-8 -*-  

import openpyxl

def main():

    try:
        wb = openpyxl.load_workbook('BugList.xlsx')
    except Exception,e:
        print str(e)

    count_S=0
    count_A=0

    for sheetName in wb.get_sheet_names():

        
        sheet=wb[sheetName]
        # print('The %s total line is %d' %(sheetName,sheet.max_row))
        

        for i in range(1,sheet.max_row):
            
            if sheet.cell('E%s'%(i)).value == 'S':
               count_S += 1
               print(sheet.cell('C%s'%(i)).value)
  
    for sheetName in wb.get_sheet_names():
        count_skip=0
        count_S=0
        count_A=0
        
        sheet=wb[sheetName]
        # print('The %s total line is %d' %(sheetName,sheet.max_row))
        
        for i in range(1,sheet.max_row):

            if sheet.cell('E%s'%(i)).value == 'A':
                count_A += 1
                print(sheet.cell('C%s'%(i)).value) 
    
    
if __name__=="__main__":
    main()
