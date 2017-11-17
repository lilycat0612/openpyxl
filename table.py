# -*- coding:utf-8 -*-  

import openpyxl

def main():

    try:
        wb = openpyxl.load_workbook('taiping.xlsx')
    except Exception,e:
        print str(e)
    
    for sheetName in wb.get_sheet_names():
        count_skip=0

        sheet=wb[sheetName]
        print('The %s total line is %d' %(sheetName,sheet.max_column))

        for i in range(1,sheet.max_column+1):
            if sheet.cell('G%s'%(i)).value == 'Planned':
                print(sheet.cell('A%s'%(i)).value) 
            else:
                count_skip
        

        #print(sheet_name)
        #print(type(sheet_name))
        #sheet=wb.get_sheet_by_name[str(sheet_name)]
        #print(sheet['A1'])
    
    
if __name__=="__main__":
    main()